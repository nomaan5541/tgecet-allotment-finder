"""
TGECET College Allotment Scraper using Selenium
Scrapes all college allotment data and saves to Excel.
Uses a real browser to handle ASP.NET postbacks properly.
"""

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import (
    StaleElementReferenceException, 
    TimeoutException,
    NoSuchElementException
)
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import sys
import os
import json
from collections import defaultdict

BASE_URL = "https://tgecet.nic.in/college_allotment.aspx"
OUTPUT_FILE = r"d:\COLLAGE BROS\TGECET_2026_College_Allotments.xlsx"
PROGRESS_FILE = r"d:\COLLAGE BROS\scrape_progress.json"


def setup_driver():
    """Setup Chrome WebDriver in headless mode."""
    options = Options()
    options.add_argument('--headless=new')
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-gpu')
    options.add_argument('--window-size=1920,1080')
    options.add_argument('--disable-blink-features=AutomationControlled')
    options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36')
    
    driver = webdriver.Chrome(options=options)
    driver.set_page_load_timeout(60)
    driver.implicitly_wait(5)
    return driver


def wait_for_page_ready(driver, timeout=20):
    """Wait for ASP.NET postback to complete."""
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        time.sleep(0.5)
    except TimeoutException:
        print("    WARNING: Page load timeout, continuing anyway...")


def navigate_to_allotment_page(driver):
    """Navigate to allotment page properly (home page first for session)."""
    driver.get("https://tgecet.nic.in/default.aspx")
    wait_for_page_ready(driver, timeout=30)
    time.sleep(2)
    driver.execute_script("window.open('https://tgecet.nic.in/college_allotment.aspx', '_self')")
    wait_for_page_ready(driver, timeout=30)
    time.sleep(1)
    # Verify page loaded
    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.ID, "MainContent_DropDownList1"))
    )



def get_college_list(driver):
    """Get list of all colleges from the dropdown."""
    college_select = Select(driver.find_element(By.ID, "MainContent_DropDownList1"))
    colleges = []
    for option in college_select.options:
        value = option.get_attribute('value')
        text = option.text.strip()
        if value and value != '0':
            colleges.append({'code': value, 'name': text})
    return colleges


def get_branch_list(driver):
    """Get list of branches from the dropdown (after a college is selected)."""
    branch_select = Select(driver.find_element(By.ID, "MainContent_DropDownList2"))
    branches = []
    for option in branch_select.options:
        value = option.get_attribute('value')
        text = option.text.strip()
        if value and value != '0':
            branches.append({'code': value, 'name': text})
    return branches


def select_college(driver, college_code):
    """Select a college from the dropdown, triggering postback."""
    try:
        college_select = Select(driver.find_element(By.ID, "MainContent_DropDownList1"))
        college_select.select_by_value(college_code)
        wait_for_page_ready(driver)
        time.sleep(0.5)
        return True
    except Exception as e:
        print(f"    Error selecting college {college_code}: {e}")
        return False


def select_branch_and_submit(driver, college_code, branch_code):
    """Select a branch and click Show Allotments."""
    try:
        # Make sure college is still selected
        college_select = Select(driver.find_element(By.ID, "MainContent_DropDownList1"))
        current_val = college_select.first_selected_option.get_attribute('value')
        if current_val != college_code:
            college_select.select_by_value(college_code)
            wait_for_page_ready(driver)
        
        # Select branch
        branch_select = Select(driver.find_element(By.ID, "MainContent_DropDownList2"))
        branch_select.select_by_value(branch_code)
        time.sleep(0.3)
        
        # Click Show Allotments button
        btn = driver.find_element(By.ID, "MainContent_btn_allot")
        btn.click()
        wait_for_page_ready(driver)
        time.sleep(0.5)
        return True
    except Exception as e:
        print(f"    Error with branch {branch_code}: {e}")
        return False


def scrape_table(driver):
    """Scrape student data from the results table."""
    students = []
    try:
        # Find the data table
        tables = driver.find_elements(By.TAG_NAME, 'table')
        data_table = None
        
        for table in tables:
            cls = table.get_attribute('class') or ''
            if 'sortable' in cls:
                data_table = table
                break
        
        if not data_table:
            # Try GridView
            try:
                data_table = driver.find_element(By.CSS_SELECTOR, 'table[id*="GridView"]')
            except NoSuchElementException:
                pass
        
        if not data_table:
            return students
        
        rows = data_table.find_elements(By.TAG_NAME, 'tr')
        
        for row in rows[1:]:  # Skip header
            cells = row.find_elements(By.TAG_NAME, 'td')
            if len(cells) >= 8:
                student = {
                    'S.No': cells[0].text.strip(),
                    'Hall Ticket No': cells[1].text.strip(),
                    'Rank': cells[2].text.strip(),
                    'Name': cells[3].text.strip(),
                    'Sex': cells[4].text.strip(),
                    'Caste': cells[5].text.strip(),
                    'Region': cells[6].text.strip(),
                    'Seat Category': cells[7].text.strip(),
                }
                students.append(student)
    
    except StaleElementReferenceException:
        print("    WARNING: Stale element, page may have changed")
    except Exception as e:
        print(f"    Error scraping table: {e}")
    
    return students


def load_progress():
    """Load progress from file to enable resume."""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, 'r') as f:
            return json.load(f)
    return {'completed_colleges': [], 'all_data': []}


def save_progress(completed_colleges, all_data):
    """Save progress to file."""
    with open(PROGRESS_FILE, 'w') as f:
        json.dump({
            'completed_colleges': completed_colleges,
            'all_data': all_data
        }, f)


def save_to_excel(all_data, filename):
    """Save all scraped data to a well-formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Allotments"
    
    # Styles
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Calibri', size=11)
    data_align = Alignment(horizontal='left', vertical='center')
    center_align = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    
    # Headers
    headers = [
        'S.No', 'College Code', 'College Name', 'Branch Code', 'Branch Name',
        'Hall Ticket No', 'Rank', 'Name of Candidate', 'Sex', 'Caste', 'Region', 'Seat Category'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    
    ws.freeze_panes = 'A2'
    
    # Data
    for idx, record in enumerate(all_data):
        row_data = [
            idx + 1,
            record.get('college_code', ''),
            record.get('college_name', ''),
            record.get('branch_code', ''),
            record.get('branch_name', ''),
            record.get('Hall Ticket No', ''),
            record.get('Rank', ''),
            record.get('Name', ''),
            record.get('Sex', ''),
            record.get('Caste', ''),
            record.get('Region', ''),
            record.get('Seat Category', ''),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 2, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = center_align if col_idx in [1, 7, 9] else data_align
            if idx % 2 == 1:
                cell.fill = alt_fill
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(all_data) + 1}"
    
    # Column widths
    col_widths = {'A': 8, 'B': 12, 'C': 55, 'D': 12, 'E': 55,
                  'F': 18, 'G': 10, 'H': 30, 'I': 8, 'J': 12, 'K': 10, 'L': 18}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # Summary sheet
    ws2 = wb.create_sheet("Summary by College-Branch")
    summary_headers = ['College Code', 'College Name', 'Branch Code', 'Branch Name', 'Total Students']
    for col_idx, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws2.freeze_panes = 'A2'
    
    counts = defaultdict(lambda: {'cn': '', 'bn': '', 'count': 0})
    for r in all_data:
        key = (r['college_code'], r['branch_code'])
        counts[key]['cn'] = r['college_name']
        counts[key]['bn'] = r['branch_name']
        counts[key]['count'] += 1
    
    row = 2
    for (cc, bc), info in sorted(counts.items()):
        for col_idx, val in enumerate([cc, info['cn'], bc, info['bn'], info['count']], 1):
            cell = ws2.cell(row=row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_align
        row += 1
    
    ws2.auto_filter.ref = f"A1:E{row - 1}"
    for col, w in {'A': 12, 'B': 55, 'C': 12, 'D': 55, 'E': 15}.items():
        ws2.column_dimensions[col].width = w
    
    wb.save(filename)
    print(f"\n{'='*60}")
    print(f"  Excel saved: {filename}")
    print(f"  Total records: {len(all_data)}")
    print(f"{'='*60}")


def main():
    print("=" * 60)
    print("  TGECET 2026 - College Allotment Scraper (Selenium)")
    print("=" * 60)
    
    # Load progress
    progress = load_progress()
    completed_colleges = set(progress['completed_colleges'])
    all_data = progress['all_data']
    
    if all_data:
        print(f"  Resuming from checkpoint: {len(all_data)} records, {len(completed_colleges)} colleges done")
    
    driver = setup_driver()
    
    try:
        # Load the page - MUST visit home page first to establish session
        print("\nLoading home page first to establish session...")
        max_load_retries = 5
        for attempt in range(max_load_retries):
            try:
                # Visit home page first (establishes sessPOLYCET cookie)
                driver.get("https://tgecet.nic.in/default.aspx")
                wait_for_page_ready(driver, timeout=30)
                time.sleep(2)
                print(f"  Home page loaded. Navigating to allotment page...")
                
                # Now navigate to allotment page via JS (driver.get doesn't work - causes 500)
                driver.execute_script("window.open('https://tgecet.nic.in/college_allotment.aspx', '_self')")
                wait_for_page_ready(driver, timeout=30)
                time.sleep(1)
                
                # Check if the college dropdown exists
                WebDriverWait(driver, 15).until(
                    EC.presence_of_element_located((By.ID, "MainContent_DropDownList1"))
                )
                
                # Verify the page title is correct (not an error page)
                title = driver.title
                if "Object reference" in title or "Error" in title:
                    raise Exception(f"Error page loaded: {title}")
                
                print("  Allotment page loaded successfully!")
                break
            except Exception as e:
                print(f"  Attempt {attempt+1} failed: {e}")
                if attempt < max_load_retries - 1:
                    time.sleep(5)
                    # Restart browser on repeated failures
                    if attempt >= 2:
                        try:
                            driver.quit()
                        except:
                            pass
                        driver = setup_driver()
                else:
                    print("FATAL: Could not load the page!")
                    return
        
        # Get all colleges
        colleges = get_college_list(driver)
        total_colleges = len(colleges)
        print(f"  Found {total_colleges} colleges")
        
        failed = []
        
        for c_idx, college in enumerate(colleges):
            cc = college['code']
            cn = college['name']
            
            if cc in completed_colleges:
                continue
            
            print(f"\n[{c_idx + 1}/{total_colleges}] {cn}")
            
            try:
                # Select college
                if not select_college(driver, cc):
                    # Retry by reloading
                    try:
                        navigate_to_allotment_page(driver)
                    except:
                        pass
                    if not select_college(driver, cc):
                        failed.append((cc, 'SELECT_FAILED'))
                        continue
                
                # Get branches
                branches = get_branch_list(driver)
                if not branches:
                    print(f"  -> No branches available")
                    completed_colleges.add(cc)
                    continue
                
                print(f"  -> {len(branches)} branches: {', '.join(b['code'] for b in branches)}")
                
                for b_idx, branch in enumerate(branches):
                    bc = branch['code']
                    bn = branch['name']
                    
                    try:
                        if select_branch_and_submit(driver, cc, bc):
                            students = scrape_table(driver)
                            
                            for s in students:
                                s['college_code'] = cc
                                s['college_name'] = cn
                                s['branch_code'] = bc
                                s['branch_name'] = bn
                            
                            all_data.extend(students)
                            print(f"     [{bc}] {len(students)} students")
                        else:
                            print(f"     [{bc}] FAILED to submit")
                            failed.append((cc, bc))
                        
                        time.sleep(0.3)
                        
                    except Exception as e:
                        print(f"     [{bc}] ERROR: {e}")
                        failed.append((cc, bc))
                        # Try to recover
                        try:
                            navigate_to_allotment_page(driver)
                            select_college(driver, cc)
                        except:
                            pass
                        time.sleep(1)
                
                completed_colleges.add(cc)
                
                # Save progress every 10 colleges
                if (c_idx + 1) % 10 == 0:
                    save_progress(list(completed_colleges), all_data)
                    print(f"\n  === Progress saved: {len(all_data)} records from {len(completed_colleges)} colleges ===\n")
                
                time.sleep(0.3)
                
            except Exception as e:
                print(f"  -> ERROR: {e}")
                failed.append((cc, 'GENERAL'))
                # Full recovery
                try:
                    navigate_to_allotment_page(driver)
                except:
                    try:
                        driver.quit()
                        driver = setup_driver()
                        navigate_to_allotment_page(driver)
                    except:
                        pass
                time.sleep(2)
        
        # Final save
        save_progress(list(completed_colleges), all_data)
        
        print(f"\n\n{'='*60}")
        print(f"  SCRAPING COMPLETE!")
        print(f"  Total students: {len(all_data)}")
        print(f"  Colleges processed: {len(completed_colleges)}/{total_colleges}")
        print(f"  Failed combos: {len(failed)}")
        print(f"{'='*60}")
        
        if all_data:
            save_to_excel(all_data, OUTPUT_FILE)
        
        if failed:
            print(f"\nFailed combinations:")
            for f in failed:
                print(f"  {f}")
    
    finally:
        driver.quit()


if __name__ == '__main__':
    main()
