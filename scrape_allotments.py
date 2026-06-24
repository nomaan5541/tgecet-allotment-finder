"""
TGECET College Allotment Scraper
Scrapes all college allotment data from https://tgecet.nic.in/college_allotment.aspx
and saves to an Excel file.
"""

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import time
import sys
import re

BASE_URL = "https://tgecet.nic.in/college_allotment.aspx"
HOME_URL = "https://tgecet.nic.in/"

def create_session():
    """Create a requests session with proper headers."""
    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Referer': HOME_URL,
    })
    return session


def get_asp_fields(soup):
    """Extract ASP.NET hidden form fields from the page."""
    fields = {}
    for field_name in ['__VIEWSTATE', '__VIEWSTATEGENERATOR', '__EVENTVALIDATION', 
                       '__EVENTTARGET', '__EVENTARGUMENT', '__LASTFOCUS']:
        field = soup.find('input', {'name': field_name})
        if field:
            fields[field_name] = field.get('value', '')
    return fields


def get_initial_page(session):
    """Load the initial page and get the college list."""
    # First visit home page to establish session
    print("Visiting home page to establish session...")
    try:
        session.get(HOME_URL, timeout=30)
    except Exception as e:
        print(f"Warning: Could not reach home page: {e}")
    
    time.sleep(1)
    
    # Now visit the allotment page
    print("Loading college allotment page...")
    resp = session.get(BASE_URL, timeout=30)
    resp.raise_for_status()
    
    soup = BeautifulSoup(resp.text, 'lxml')
    
    # Get college dropdown options
    college_select = soup.find('select', {'id': 'MainContent_DropDownList1'})
    if not college_select:
        print("ERROR: Could not find college dropdown!")
        print("Page content preview:", resp.text[:500])
        sys.exit(1)
    
    colleges = []
    for option in college_select.find_all('option'):
        value = option.get('value', '')
        text = option.get_text(strip=True)
        if value and value != '0':  # Skip the default "Select" option
            colleges.append({'code': value, 'name': text})
    
    asp_fields = get_asp_fields(soup)
    
    print(f"Found {len(colleges)} colleges")
    return colleges, asp_fields, soup


def get_branches_for_college(session, college_code, asp_fields):
    """Select a college and get the available branches via ASP.NET postback."""
    form_data = {
        '__EVENTTARGET': 'SMPage$MainContent$DropDownList1',
        '__EVENTARGUMENT': '',
        '__LASTFOCUS': '',
        '__VIEWSTATE': asp_fields['__VIEWSTATE'],
        '__VIEWSTATEGENERATOR': asp_fields.get('__VIEWSTATEGENERATOR', ''),
        '__EVENTVALIDATION': asp_fields.get('__EVENTVALIDATION', ''),
        'SMPage$MainContent$DropDownList1': college_code,
        'SMPage$MainContent$DropDownList2': '0',
    }
    
    resp = session.post(BASE_URL, data=form_data, timeout=30)
    resp.raise_for_status()
    
    soup = BeautifulSoup(resp.text, 'lxml')
    
    # Get branch dropdown options
    branch_select = soup.find('select', {'id': 'MainContent_DropDownList2'})
    branches = []
    if branch_select:
        for option in branch_select.find_all('option'):
            value = option.get('value', '')
            text = option.get_text(strip=True)
            if value and value != '0':  # Skip default
                branches.append({'code': value, 'name': text})
    
    new_asp_fields = get_asp_fields(soup)
    
    return branches, new_asp_fields


def get_allotment_data(session, college_code, branch_code, asp_fields):
    """Submit the form to get allotment data for a specific college-branch."""
    form_data = {
        '__EVENTTARGET': '',
        '__EVENTARGUMENT': '',
        '__LASTFOCUS': '',
        '__VIEWSTATE': asp_fields['__VIEWSTATE'],
        '__VIEWSTATEGENERATOR': asp_fields.get('__VIEWSTATEGENERATOR', ''),
        '__EVENTVALIDATION': asp_fields.get('__EVENTVALIDATION', ''),
        'SMPage$MainContent$DropDownList1': college_code,
        'SMPage$MainContent$DropDownList2': branch_code,
        'SMPage$MainContent$btn_allot': 'Show Allotments',
    }
    
    resp = session.post(BASE_URL, data=form_data, timeout=30)
    resp.raise_for_status()
    
    soup = BeautifulSoup(resp.text, 'lxml')
    
    # Find the results table
    table = soup.find('table', class_='sortable')
    students = []
    
    if table:
        rows = table.find_all('tr')
        for row in rows[1:]:  # Skip header row
            cells = row.find_all('td')
            if len(cells) >= 8:
                student = {
                    'S.No': cells[0].get_text(strip=True),
                    'Hall Ticket No': cells[1].get_text(strip=True),
                    'Rank': cells[2].get_text(strip=True),
                    'Name': cells[3].get_text(strip=True),
                    'Sex': cells[4].get_text(strip=True),
                    'Caste': cells[5].get_text(strip=True),
                    'Region': cells[6].get_text(strip=True),
                    'Seat Category': cells[7].get_text(strip=True),
                }
                students.append(student)
    
    new_asp_fields = get_asp_fields(soup)
    
    return students, new_asp_fields


def save_to_excel(all_data, filename="TGECET_2026_College_Allotments.xlsx"):
    """Save all scraped data to a well-formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Allotments"
    
    # Styling
    header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Calibri', size=11)
    data_alignment = Alignment(horizontal='left', vertical='center')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    
    # Headers
    headers = [
        'S.No', 'College Code', 'College Name', 'Branch Code', 'Branch Name',
        'Hall Ticket No', 'Rank', 'Name of Candidate', 'Sex', 'Caste', 'Region', 'Seat Category'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Freeze the header row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:L{len(all_data) + 1}"
    
    # Data rows
    row_idx = 2
    for idx, record in enumerate(all_data):
        row_data = [
            idx + 1,
            record.get('college_code', ''),
            record.get('college_name', ''),
            record.get('branch_code', ''),
            record.get('branch_name', ''),
            record.get('Hall Ticket No', ''),
            record.get('Rank', ''),
            record.get('Name', ''),
            record.get('Sex', ''),
            record.get('Caste', ''),
            record.get('Region', ''),
            record.get('Seat Category', ''),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            if col_idx in [1, 7, 9]:  # S.No, Rank, Sex - center align
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
            
            # Alternate row coloring
            if idx % 2 == 1:
                cell.fill = alt_fill
        
        row_idx += 1
    
    # Column widths
    col_widths = {
        'A': 8,   # S.No
        'B': 12,  # College Code
        'C': 55,  # College Name
        'D': 12,  # Branch Code
        'E': 55,  # Branch Name
        'F': 18,  # Hall Ticket No
        'G': 10,  # Rank
        'H': 30,  # Name
        'I': 8,   # Sex
        'J': 12,  # Caste
        'K': 10,  # Region
        'L': 18,  # Seat Category
    }
    
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # ===== Create a Summary Sheet =====
    ws_summary = wb.create_sheet("Summary by College")
    
    # Summary headers
    summary_headers = ['College Code', 'College Name', 'Branch Code', 'Branch Name', 'Total Students']
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws_summary.freeze_panes = 'A2'
    
    # Aggregate data
    from collections import defaultdict
    college_branch_counts = defaultdict(lambda: {'college_name': '', 'branch_name': '', 'count': 0})
    
    for record in all_data:
        key = (record['college_code'], record['branch_code'])
        college_branch_counts[key]['college_name'] = record['college_name']
        college_branch_counts[key]['branch_name'] = record['branch_name']
        college_branch_counts[key]['count'] += 1
    
    summary_row = 2
    for (cc, bc), info in sorted(college_branch_counts.items()):
        row_data = [cc, info['college_name'], bc, info['branch_name'], info['count']]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=summary_row, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border
            cell.alignment = data_alignment
            if (summary_row - 2) % 2 == 1:
                cell.fill = alt_fill
        summary_row += 1
    
    ws_summary.auto_filter.ref = f"A1:E{summary_row - 1}"
    
    summary_col_widths = {'A': 12, 'B': 55, 'C': 12, 'D': 55, 'E': 15}
    for col_letter, width in summary_col_widths.items():
        ws_summary.column_dimensions[col_letter].width = width
    
    # Save
    wb.save(filename)
    print(f"\n{'='*60}")
    print(f"Excel file saved: {filename}")
    print(f"Total records: {len(all_data)}")
    print(f"{'='*60}")


def main():
    print("=" * 60)
    print("  TGECET 2026 - College Allotment Scraper")
    print("=" * 60)
    
    session = create_session()
    
    # Step 1: Get all colleges
    colleges, asp_fields, soup = get_initial_page(session)
    
    all_data = []
    total_colleges = len(colleges)
    failed_combos = []
    
    for c_idx, college in enumerate(colleges):
        college_code = college['code']
        college_name = college['name']
        
        print(f"\n[{c_idx + 1}/{total_colleges}] Processing: {college_name}")
        
        try:
            # Step 2: Get branches for this college
            branches, asp_fields = get_branches_for_college(session, college_code, asp_fields)
            
            if not branches:
                print(f"  -> No branches found for {college_code}, skipping")
                continue
            
            print(f"  -> Found {len(branches)} branches: {', '.join(b['code'] for b in branches)}")
            
            for b_idx, branch in enumerate(branches):
                branch_code = branch['code']
                branch_name = branch['name']
                
                try:
                    students, asp_fields = get_allotment_data(
                        session, college_code, branch_code, asp_fields
                    )
                    
                    for student in students:
                        student['college_code'] = college_code
                        student['college_name'] = college_name
                        student['branch_code'] = branch_code
                        student['branch_name'] = branch_name
                    
                    all_data.extend(students)
                    print(f"     [{branch_code}] {len(students)} students")
                    
                    # Small delay to be polite to the server
                    time.sleep(0.3)
                    
                except Exception as e:
                    print(f"     [{branch_code}] ERROR: {e}")
                    failed_combos.append((college_code, branch_code, str(e)))
                    # Try to recover by re-loading the page
                    try:
                        _, asp_fields = get_branches_for_college(session, college_code, asp_fields)
                    except:
                        # If recovery fails, re-initialize
                        try:
                            _, asp_fields, _ = get_initial_page(session)
                        except:
                            pass
                    time.sleep(1)
            
            # Small delay between colleges
            time.sleep(0.3)
            
        except Exception as e:
            print(f"  -> ERROR processing college {college_code}: {e}")
            failed_combos.append((college_code, 'ALL', str(e)))
            # Try to recover
            try:
                _, asp_fields, _ = get_initial_page(session)
            except Exception as re_e:
                print(f"  -> Recovery failed: {re_e}")
                time.sleep(5)
                try:
                    session = create_session()
                    _, asp_fields, _ = get_initial_page(session)
                except:
                    pass
            time.sleep(1)
    
    # Save results
    print(f"\n\nTotal students scraped: {len(all_data)}")
    
    if all_data:
        output_file = r"d:\COLLAGE BROS\TGECET_2026_College_Allotments.xlsx"
        save_to_excel(all_data, output_file)
    else:
        print("No data was scraped!")
    
    if failed_combos:
        print(f"\nFailed combinations ({len(failed_combos)}):")
        for fc in failed_combos:
            print(f"  College: {fc[0]}, Branch: {fc[1]}, Error: {fc[2]}")


if __name__ == '__main__':
    main()
