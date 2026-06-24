"""
Convert the scraped data from progress JSON to a clean JSON file for the web app.
Also regenerates the Excel with any retry data.
"""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict

PROGRESS_FILE = r"d:\COLLAGE BROS\scrape_progress.json"
JSON_OUTPUT = r"d:\COLLAGE BROS\allotment_data.json"
EXCEL_OUTPUT = r"d:\COLLAGE BROS\TGECET_2026_College_Allotments.xlsx"

# Load progress data
with open(PROGRESS_FILE, 'r') as f:
    progress = json.load(f)

all_data = progress['all_data']
print(f"Total records: {len(all_data)}")

# Deduplicate by Hall Ticket No + Branch Code
seen = set()
unique_data = []
for record in all_data:
    key = (record.get('Hall Ticket No', ''), record.get('college_code', ''), record.get('branch_code', ''))
    if key not in seen and record.get('Hall Ticket No'):
        seen.add(key)
        unique_data.append(record)

print(f"Unique records (after dedup): {len(unique_data)}")

# Save clean JSON for web app
web_data = []
for r in unique_data:
    web_data.append({
        'sno': r.get('S.No', ''),
        'ht': r.get('Hall Ticket No', ''),
        'rank': r.get('Rank', ''),
        'name': r.get('Name', ''),
        'sex': r.get('Sex', ''),
        'caste': r.get('Caste', ''),
        'region': r.get('Region', ''),
        'seat_cat': r.get('Seat Category', ''),
        'cc': r.get('college_code', ''),
        'cn': r.get('college_name', ''),
        'bc': r.get('branch_code', ''),
        'bn': r.get('branch_name', ''),
    })

with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
    json.dump(web_data, f, ensure_ascii=False)

print(f"JSON saved: {JSON_OUTPUT} ({len(web_data)} records)")

# Regenerate Excel
print("\nRegenerating Excel...")
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "All Allotments"

header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_font = Font(name='Calibri', size=11)
data_align = Alignment(horizontal='left', vertical='center')
center_align = Alignment(horizontal='center', vertical='center')
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
alt_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

headers = ['S.No', 'College Code', 'College Name', 'Branch Code', 'Branch Name',
           'Hall Ticket No', 'Rank', 'Name of Candidate', 'Sex', 'Caste', 'Region', 'Seat Category']

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.freeze_panes = 'A2'

for idx, record in enumerate(unique_data):
    row_data = [idx + 1, record.get('college_code', ''), record.get('college_name', ''),
                record.get('branch_code', ''), record.get('branch_name', ''),
                record.get('Hall Ticket No', ''), record.get('Rank', ''),
                record.get('Name', ''), record.get('Sex', ''), record.get('Caste', ''),
                record.get('Region', ''), record.get('Seat Category', '')]
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=idx + 2, column=col_idx, value=value)
        cell.font = data_font
        cell.border = thin_border
        cell.alignment = center_align if col_idx in [1, 7, 9] else data_align
        if idx % 2 == 1:
            cell.fill = alt_fill

ws.auto_filter.ref = f"A1:L{len(unique_data) + 1}"
for col, w in {'A': 8, 'B': 12, 'C': 55, 'D': 12, 'E': 55, 'F': 18, 'G': 10, 'H': 30, 'I': 8, 'J': 12, 'K': 10, 'L': 18}.items():
    ws.column_dimensions[col].width = w

# Summary sheet
ws2 = wb.create_sheet("Summary")
summary_headers = ['College Code', 'College Name', 'Branch Code', 'Branch Name', 'Total Students']
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
ws2.freeze_panes = 'A2'

counts = defaultdict(lambda: {'cn': '', 'bn': '', 'count': 0})
for r in unique_data:
    key = (r['college_code'], r['branch_code'])
    counts[key]['cn'] = r['college_name']
    counts[key]['bn'] = r['branch_name']
    counts[key]['count'] += 1

row = 2
for (cc, bc), info in sorted(counts.items()):
    for col_idx, val in enumerate([cc, info['cn'], bc, info['bn'], info['count']], 1):
        cell = ws2.cell(row=row, column=col_idx, value=val)
        cell.font = data_font
        cell.border = thin_border
    row += 1

ws2.auto_filter.ref = f"A1:E{row - 1}"
for col, w in {'A': 12, 'B': 55, 'C': 12, 'D': 55, 'E': 15}.items():
    ws2.column_dimensions[col].width = w

wb.save(EXCEL_OUTPUT)
print(f"Excel saved: {EXCEL_OUTPUT}")
print("Done!")
